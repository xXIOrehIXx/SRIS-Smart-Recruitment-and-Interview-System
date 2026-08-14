"""
============================================================================
 GỘP MỌI SỐ ĐO VÀO MỘT FILE EXCEL

 Đọc: out/may_cham_4_ban.csv (tầng máy) + out/<ver>/nguoi_cham_tong_ket.csv,
      nguoi_cham_tung_dong.csv, nguoi_cham_bo_sot.csv (tầng người) -> out/KET_QUA_TONG_HOP.xlsx

 Sáu tab:
   1. TongHop   — bảng đầu bài: mỗi phiên bản một dòng, máy và người cạnh nhau
   2. TheoTin   — 10 tin x 4 phiên bản, xem tin nào kéo tụt chỉ số
   3-5. Ba bảng chấm tay bê nguyên từ out/<ver>/*.csv

 Chạy:  ../../ai-service/.venv/Scripts/python.exe 4_gop_ket_qua_excel.py
============================================================================
"""

import csv
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = Path(__file__).parent
OUT = HERE / "out"
VERS = ["v1", "v2", "v3", "v4"]

MO_TA_VER = {
    "v1": "Câu lệnh trần — không luật, không ví dụ, không ép JSON",
    "v2": "+ Ràng buộc JSON schema (Pydantic) & temperature=0",
    "v3": "+ Luật nghiệp vụ (yêu cầu vs đầu việc, bỏ giấy tờ, tách, trần 10)",
    "v4": "+ Khối ví dụ mẫu (few-shot) = PROMPT ĐANG CHẠY THẬT",
}

TEN_TIN = {
    "J01_ke_toan": "Kế toán tổng hợp",
    "J02_kinh_doanh": "Nhân viên kinh doanh B2B",
    "J03_kho_van": "Nhân viên kho",
    "J04_dotnet": "Lập trình viên .NET",
    "J05_hanh_chinh": "Nhân viên hành chính nhân sự",
    "J06_le_tan": "Lễ tân khách sạn",
    "J07_marketing": "Nhân viên Marketing",
    "J09_chi_dau_viec": "Công nhân SX — tin CHỈ CÓ đầu việc",
    "J10_qua_nhieu": "Trưởng nhóm PM — tin QUÁ DÀI (13 yêu cầu)",
    "J08_cskh": "Nhân viên chăm sóc khách hàng",
}

NHAN_MO_TA = {
    "DUNG": "dùng được",
    "BIA": "bịa — không có căn cứ trong tin",
    "DAUVIEC": "đầu việc — không phải yêu cầu ứng viên",
    "GIAYTO": "giấy tờ — đọc hồ sơ là biết",
    "GOP": "gộp — nhiều kỹ năng một dòng",
    "TRUNG": "trùng — đã có dòng khác nói cùng chuyện",
}

XANH = PatternFill("solid", fgColor="1F4E79")
XAM = PatternFill("solid", fgColor="DDEBF7")
VANG = PatternFill("solid", fgColor="FFF2CC")
LUC = PatternFill("solid", fgColor="E2EFDA")
DO = PatternFill("solid", fgColor="FCE4E4")
TRANG = Font(color="FFFFFF", bold=True)
VIEN = Border(*[Side("thin", color="B0B0B0")] * 4)


def doc(path):
    with Path(path).open(encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def so(v, mac_dinh=None):
    try:
        return float(v)
    except (TypeError, ValueError):
        return mac_dinh


def dat_header(ws, hang, tieu_de):
    for i, t in enumerate(tieu_de, start=1):
        if t == "":          # ô đã bị merge từ hàng trên — ghi vào là AttributeError
            continue
        c = ws.cell(row=hang, column=i, value=t)
        c.fill, c.font = XANH, TRANG
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = VIEN


def rong(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def tieu_de_to(ws, hang, text, span):
    c = ws.cell(row=hang, column=1, value=text)
    c.font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(start_row=hang, start_column=1, end_row=hang, end_column=span)


# ---------------------------------------------------------------- tab 1
def tab_huong_dan(wb):
    """Tab đầu tiên: ai chấm cái gì, file nào ra số nào. Mở workbook là thấy ngay."""
    ws = wb.create_sheet("DocTruoc")
    rong(ws, [4, 34, 12, 62])

    tieu_de_to(ws, 1, "ĐỌC TRƯỚC — thí nghiệm này đo cái gì, ai chấm, file nào ra số nào", 4)
    ws.cell(row=2, column=2,
            value="AI bóc tiêu chí từ tin tuyển dụng · 10 tin đa ngành × 2 lượt × 4 bậc prompt · qwen2.5 7B, temperature=0")

    h = 4
    ws.cell(row=h, column=2, value="HAI TẦNG ĐO — vì sao cần cả hai").font = Font(bold=True, size=11, color="1F4E79")
    h += 1
    dat_header(ws, h, ["", "Tầng", "Ai chấm", "Trả lời câu hỏi gì · điểm mù"])
    for tang, ai, mo in [
        ("Tầng 1 — máy đo", "🤖 máy (may_cham.py)",
         "Đếm được: dòng giấy tờ, dòng gộp, dòng trùng, vượt trần, hai lượt có giống nhau không, giây/tin. "
         "ĐIỂM MÙ: không biết một tiêu chí có DÙNG ĐƯỢC hay không."),
        ("Tầng 2 — người chấm", "🧑 người (LUAT_NGUOI_CHAM.md)",
         "Gán 1 trong 6 nhãn cho từng tiêu chí → ra Precision / Recall / F1. "
         "ĐIỂM MÙ: không lặp lại được, không đo nổi độ ổn định giữa hai lượt chạy."),
    ]:
        h += 1
        for i, g in enumerate(["", tang, ai, mo], start=1):
            if g == "":
                continue
            c = ws.cell(row=h, column=i, value=g)
            c.border, c.alignment = VIEN, Alignment(wrap_text=True, vertical="top")
            if i == 2:
                c.font = Font(bold=True)
        ws.row_dimensions[h].height = 58

    h += 2
    ws.cell(row=h, column=2, value="BA TAB CÒN LẠI").font = Font(bold=True, size=11, color="1F4E79")
    h += 1
    dat_header(ws, h, ["", "Tab", "Tầng", "Nội dung"])
    for ten, tang, mo in [
        ("TongHop", "🤖+🧑", "Bảng đầu bài: mỗi bậc prompt một dòng, số máy và số người cạnh nhau, kèm ngưỡng đánh giá."),
        ("TheoTin", "🧑", "10 tin × 4 bậc dạng bảng ngang — xem tin nào kéo tụt chỉ số. J09 (tin chỉ có đầu việc) và J10 (tin quá dài) là hai ca đối chứng."),
        ("NguoiCham_TungDong", "🧑", "Nguồn: out/<ver>/nguoi_cham_tung_dong.csv · 299 dòng, mỗi tiêu chí một nhãn + lý do chấm. Lọc được theo bậc và theo nhãn."),
        ("NguoiCham_BoSot", "🧑", "Nguồn: out/<ver>/nguoi_cham_bo_sot.csv · thứ AI QUÊN không nêu — mẫu số của recall, không đếm thì recall luôn bằng 1."),
        ("NguoiCham_TongKet", "🧑", "Nguồn: out/<ver>/nguoi_cham_tong_ket.csv · P/R/F1 từng tin + dòng TỔNG của mỗi bậc, kèm phân rã 5 kiểu lỗi."),
    ]:
        h += 1
        for i, g in enumerate(["", ten, tang, mo], start=1):
            if g == "":
                continue
            c = ws.cell(row=h, column=i, value=g)
            c.border, c.alignment = VIEN, Alignment(wrap_text=True, vertical="top")
            if i == 2:
                c.font = Font(bold=True)
        ws.row_dimensions[h].height = 30

    h += 2
    ws.cell(row=h, column=2, value="MUỐN CHẤM LẠI / SỬA NHÃN").font = Font(bold=True, size=11, color="1F4E79")
    for dong in [
        "Nhãn nằm trong 2_nguoi_cham_dien_nhan.py — KHÔNG sửa tay vào out/<ver>/nguoi_cham_tung_dong.csv (chạy lại script 1 là mất).",
        "Sửa xong chạy:  2_nguoi_cham_dien_nhan.py  →  3_nguoi_cham_tinh_diem.py --tag v4  →  4_gop_ket_qua_excel.py",
        "Luật chấm (6 nhãn + các ca ranh giới đã chốt sẵn): LUAT_NGUOI_CHAM.md",
        "Bản tường thuật đầy đủ kèm ví dụ thật: out/KET_QUA.md",
    ]:
        h += 1
        ws.cell(row=h, column=2, value="• " + dong)
        ws.merge_cells(start_row=h, start_column=2, end_row=h, end_column=4)

    h += 2
    c = ws.cell(row=h, column=2,
                value="⚠ Nhãn tầng người là bản do trợ lý AI soạn theo LUAT_NGUOI_CHAM.md, người làm đề tài rà lại. "
                      "Phải nói rõ điều này khi trích số — xem mục 7 của out/KET_QUA.md.")
    c.fill, c.alignment = VANG, Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=h, start_column=2, end_row=h, end_column=4)
    ws.row_dimensions[h].height = 34


def tab_tong_hop(wb, may, nguoi):
    ws = wb.create_sheet("TongHop")
    rong(ws, [8, 58, 13, 13, 12, 12, 12, 12, 12, 12, 11, 11])

    tieu_de_to(ws, 1, "AI ĐỀ XUẤT TIÊU CHÍ — TỔNG HỢP 4 BẬC PROMPT (ablation)", 12)
    ws.cell(row=2, column=1,
            value="10 tin tuyển dụng đa ngành × 2 lượt × 4 phiên bản = 80 lượt gọi model · qwen2.5 7B qua Ollama · temperature=0")
    ws.cell(row=3, column=1,
            value="Tầng máy = đo tự động (may_cham.py) · Tầng người = chấm tay theo LUAT_NGUOI_CHAM.md — nhãn do trợ lý soạn, người làm đề tài rà lại")

    dat_header(ws, 5, [
        "Bản", "Thêm gì so với bậc dưới", "Tiêu chí đề xuất", "Dùng được (DUNG)",
        "Bỏ sót", "Precision", "Recall", "F1",
        "Giấy tờ %", "Gộp %", "Ổn định", "Giây/tin",
    ])
    ws.row_dimensions[5].height = 32

    h = 6
    for v in VERS:
        m = next(r for r in may if r["version"] == v)
        t = next(r for r in nguoi[v] if r["id"] == "TONG")
        gia_tri = [
            v.upper(), MO_TA_VER[v], int(t["de_xuat"]), int(t["dung"]), int(t["bo_sot"]),
            so(t["precision"]), so(t["recall"]), so(t["f1"]),
            so(m["giay_to_rate"]), so(m["gop_rate"]), so(m["on_dinh"]), so(m["giay_tb"]),
        ]
        for i, g in enumerate(gia_tri, start=1):
            c = ws.cell(row=h, column=i, value=g)
            c.border = VIEN
            c.alignment = Alignment(vertical="center", wrap_text=(i == 2))
            if i >= 6:
                c.number_format = "0.000" if i <= 8 or i == 11 else "0.0%" if i <= 10 else "0.00"
            if v == "v4":
                c.fill = LUC
                c.font = Font(bold=True)
        ws.row_dimensions[h].height = 30
        h += 1

    h += 1
    ws.cell(row=h, column=1, value="Ngưỡng đánh giá (mượn của nhóm capstone khác — AI_TESTING_REFERENCE.md)").font = Font(bold=True)
    h += 1
    dat_header(ws, h, ["", "Chỉ số", "Tốt", "Chấp nhận được", "Cần cải thiện", "V4 đạt mức"])
    for chi_so, val in [("Precision", 0.846), ("Recall", 0.873), ("F1", 0.859)]:
        h += 1
        muc = "Tốt" if val >= 0.85 else "Chấp nhận được" if val >= 0.70 else "Cần cải thiện"
        for i, g in enumerate(["", chi_so, "≥ 0.85", "0.70 – 0.84", "< 0.70", f"{val:.3f} — {muc}"], start=1):
            c = ws.cell(row=h, column=i, value=g)
            c.border = VIEN
            if i == 6:
                c.fill = LUC if muc == "Tốt" else VANG

    h += 2
    for dong in [
        "ĐỌC BẢNG NÀY THẾ NÀO:",
        "• V1→V2 (ép định dạng): precision gần như đứng yên (.593→.608). Cái ăn là ổn định 0,983→1,000 và cưỡng chế trần 10 ở J10.",
        "• V2→V3 (luật nghiệp vụ): precision nhảy .608→.800 — luật quét sạch nhóm 'đầu việc' (11 dòng → 0). Nhưng recall tụt .980→.842 vì cắt quá tay.",
        "• V3→V4 (ví dụ mẫu): kéo lại cả hai — precision .800→.846 VÀ recall .842→.873. Đây là bậc duy nhất cải thiện cả hai chiều.",
        "• Recall đắt hơn precision trong bài toán này: dòng thừa người duyệt xoá được, dòng thiếu thì không ai nhìn thấy (LUAT_NGUOI_CHAM.md).",
    ]:
        ws.cell(row=h, column=1, value=dong)
        h += 1

    ws.freeze_panes = "A6"


# ---------------------------------------------------------------- tab 2
def tab_theo_tin(wb, nguoi, may_theo_tin):
    ws = wb.create_sheet("TheoTin")
    rong(ws, [20, 34] + [11, 9, 8, 10, 10, 9] * 4)

    tieu_de_to(ws, 1, "TỪNG TIN QUA 4 BẬC — tin nào kéo tụt chỉ số", 26)

    hang = 3
    ws.cell(row=hang, column=1, value="Mã tin").fill = XANH
    ws.cell(row=hang, column=2, value="Vị trí").fill = XANH
    for c in (1, 2):
        ws.cell(row=hang, column=c).font = TRANG
        ws.merge_cells(start_row=hang, start_column=c, end_row=hang + 1, end_column=c)
        ws.cell(row=hang, column=c).alignment = Alignment(horizontal="center", vertical="center")
    for k, v in enumerate(VERS):
        c0 = 3 + k * 6
        c = ws.cell(row=hang, column=c0, value=v.upper())
        c.fill, c.font = XANH, TRANG
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=hang, start_column=c0, end_row=hang, end_column=c0 + 5)
    dat_header(ws, hang + 1, ["", ""] + ["Đề xuất", "DUNG", "Sót", "Precision", "Recall", "F1"] * 4)
    ws.row_dimensions[hang + 1].height = 28

    h = hang + 2
    for tin in sorted(TEN_TIN):
        ws.cell(row=h, column=1, value=tin).border = VIEN
        c = ws.cell(row=h, column=2, value=TEN_TIN[tin])
        c.border, c.alignment = VIEN, Alignment(wrap_text=True, vertical="center")
        if tin in ("J09_chi_dau_viec", "J10_qua_nhieu"):
            for cc in (1, 2):
                ws.cell(row=h, column=cc).fill = VANG
        for k, v in enumerate(VERS):
            r = next((x for x in nguoi[v] if x["id"] == tin), None)
            c0 = 3 + k * 6
            if r is None:
                continue
            gia_tri = [int(r["de_xuat"]), int(r["dung"]), int(r["bo_sot"]),
                       so(r["precision"]), so(r["recall"]), so(r["f1"])]
            for i, g in enumerate(gia_tri):
                cell = ws.cell(row=h, column=c0 + i, value=g if g is not None else "—")
                cell.border = VIEN
                cell.alignment = Alignment(horizontal="center")
                if i >= 3 and g is not None:
                    cell.number_format = "0.000"
                    if i == 5:
                        cell.fill = LUC if g >= 0.85 else DO if g < 0.70 else VANG
            if r["de_xuat"] == "0" and r["bo_sot"] == "0":
                ws.cell(row=h, column=c0 + 5, value="ĐẠT").fill = LUC
        ws.row_dimensions[h].height = 26
        h += 1

    h += 1
    for dong in [
        "J09 (tin chỉ có đầu việc): V1/V2 đẻ 7 tiêu chí RÁC từ mô tả công việc · V3/V4 trả RỖNG = đúng thiết kế, cột F1 ghi ĐẠT.",
        "Tin J08_cskh (chăm sóc khách hàng) THAY cho tin tài xế J08_tai_xe ngày 14/08/2026 — xem mục Hạn chế trong out/KET_QUA.md.",
        "J10 (13 yêu cầu, trần 10): V1 đẻ đủ 13 → vượt trần. V2/V3/V4 cắt còn 10 và bỏ ĐÚNG 3 dòng giấy tờ (PMP, fintech, bằng ĐH) → không tính là bỏ sót.",
    ]:
        ws.cell(row=h, column=1, value=dong)
        h += 1

    ws.freeze_panes = "C5"


# ---------------------------------------------------------------- tab 3
def tab_tung_dong(wb, tung_dong):
    """Nguyên văn out/<ver>/nguoi_cham_tung_dong.csv của cả 4 bậc, gộp một bảng."""
    ws = wb.create_sheet("NguoiCham_TungDong")
    rong(ws, [7, 20, 30, 62, 12, 52])

    tieu_de_to(ws, 1, "PHIẾU CHẤM TAY — MỖI TIÊU CHÍ MỘT DÒNG  (nguồn: out/<ver>/nguoi_cham_tung_dong.csv)", 6)
    ws.cell(row=2, column=1,
            value="Lọc cột 'Bậc' để xem riêng một phiên bản · lọc cột 'Nhãn' để xem riêng một kiểu lỗi. "
                  "Sửa nhãn thì sửa trong 2_nguoi_cham_dien_nhan.py rồi chạy lại, đừng sửa ở đây.")
    dat_header(ws, 4, ["Bậc", "Mã tin", "Vị trí", "Tiêu chí AI đề xuất", "Nhãn", "Ghi chú / lý do chấm"])
    ws.row_dimensions[4].height = 26

    h = 5
    for ver, r in tung_dong:
        gia_tri = [ver.upper(), r["id"], TEN_TIN[r["id"]], r["tieu_chi"], r["nhan"], r["ghi_chu"]]
        for i, g in enumerate(gia_tri, start=1):
            c = ws.cell(row=h, column=i, value=g)
            c.border = VIEN
            c.alignment = Alignment(wrap_text=i in (4, 6), vertical="center",
                                    horizontal="center" if i in (1, 5) else "left")
            if i == 5:
                c.fill = LUC if g == "DUNG" else DO
                c.font = Font(bold=True)
        ws.row_dimensions[h].height = 24
        h += 1

    ws.auto_filter.ref = f"A4:F{h - 1}"
    ws.freeze_panes = "A5"

    h += 1
    ws.cell(row=h, column=1, value="Ý nghĩa nhãn (LUAT_NGUOI_CHAM.md):").font = Font(bold=True)
    for ma, mo in NHAN_MO_TA.items():
        h += 1
        ws.cell(row=h, column=2, value=ma).font = Font(bold=True)
        ws.cell(row=h, column=4, value=mo)


# ---------------------------------------------------------------- tab 4
def tab_bo_sot(wb, bo_sot):
    """Nguyên văn out/<ver>/nguoi_cham_bo_sot.csv — mẫu số của recall."""
    ws = wb.create_sheet("NguoiCham_BoSot")
    rong(ws, [7, 20, 34, 12, 78])

    tieu_de_to(ws, 1, "BỎ SÓT — TIÊU CHÍ ĐÁNG LẼ PHẢI CÓ MÀ AI KHÔNG NÊU  (nguồn: out/<ver>/nguoi_cham_bo_sot.csv)", 5)
    for i, dong in enumerate([
        "Tab bên cạnh chỉ chấm được thứ AI ĐÃ nói ra. Thứ nó quên thì không có dòng nào để chấm — phải đọc lại tin gốc mà đếm.",
        "Recall = DUNG / (DUNG + bỏ sót). Không có cột này thì recall vĩnh viễn bằng 1, và một model chỉ nói 1 tiêu chí chắc ăn sẽ đạt precision 1.0.",
        "Chỉ đếm thứ tự nó cũng đạt chuẩn DUNG: 'Tốt nghiệp Đại học' biến mất KHÔNG tính là sót, vì bỏ nó mới đúng. J10 cắt cho đủ trần 10 cũng không tính.",
    ]):
        ws.cell(row=2 + i, column=1, value="• " + dong)

    dat_header(ws, 6, ["Bậc", "Mã tin", "Vị trí", "Số bỏ sót", "Sót cái gì"])
    ws.row_dimensions[6].height = 26

    h = 7
    for ver, r in bo_sot:
        n = int(r["so_bo_sot"] or 0)
        for i, g in enumerate([ver.upper(), r["id"], TEN_TIN[r["id"]], n, r["ghi_chu"]], start=1):
            c = ws.cell(row=h, column=i, value=g)
            c.border = VIEN
            c.alignment = Alignment(wrap_text=(i == 5), vertical="center",
                                    horizontal="center" if i in (1, 4) else "left")
            if i == 4 and n:
                c.fill = DO if n >= 2 else VANG
                c.font = Font(bold=True)
        ws.row_dimensions[h].height = 24
        h += 1

    ws.auto_filter.ref = f"A6:E{h - 1}"
    ws.freeze_panes = "A7"


# ---------------------------------------------------------------- tab 5
def tab_tong_ket(wb, nguoi):
    """Nguyên văn out/<ver>/nguoi_cham_tong_ket.csv — do 3_nguoi_cham_tinh_diem.py ghi ra."""
    ws = wb.create_sheet("NguoiCham_TongKet")
    rong(ws, [7, 22, 34, 10, 9, 8, 8, 10, 9, 8, 8, 11, 11, 10])

    tieu_de_to(ws, 1, "TỔNG KẾT TẦNG NGƯỜI — P / R / F1 TỪNG TIN  (nguồn: out/<ver>/nguoi_cham_tong_ket.csv)", 14)
    ws.cell(row=2, column=1,
            value="Dòng TỔNG của mỗi bậc là con số đem trích vào báo cáo. Ô '—' nghĩa là không áp dụng: "
                  "tin J09 không nêu yêu cầu nào nên trả rỗng là ĐÚNG, không có gì để tính tỉ lệ.")

    dat_header(ws, 4, ["Bậc", "Mã tin", "Vị trí", "Đề xuất", "DUNG", "Sót",
                       "BIA", "DAUVIEC", "GIAYTO", "GOP", "TRUNG",
                       "Precision", "Recall", "F1"])
    ws.row_dimensions[4].height = 30

    h = 5
    for ver, r in nguoi:
        la_tong = r["id"] == "TONG"
        ten = "— TỔNG CẢ BẬC —" if la_tong else TEN_TIN.get(r["id"], "")
        gia_tri = [ver.upper(), r["id"], ten,
                   int(r["de_xuat"]), int(r["dung"]), int(r["bo_sot"]),
                   int(r["bia"]), int(r["dauviec"]), int(r["giayto"]), int(r["gop"]), int(r["trung"]),
                   so(r["precision"]), so(r["recall"]), so(r["f1"])]
        for i, g in enumerate(gia_tri, start=1):
            c = ws.cell(row=h, column=i, value="—" if g is None else g)
            c.border = VIEN
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if i in (2, 3) else "center")
            if i >= 12 and g is not None:
                c.number_format = "0.000"
                if i == 14:
                    c.fill = LUC if g >= 0.85 else DO if g < 0.70 else VANG
            if la_tong:
                c.font = Font(bold=True)
                if i <= 11:
                    c.fill = XAM
        h += 1

    ws.auto_filter.ref = f"A4:N{h - 1}"
    ws.freeze_panes = "D5"


def chan_cong_thuc(wb):
    """
    Ô nào có chuỗi bắt đầu bằng '=' thì openpyxl ghi thành CÔNG THỨC, Excel parse
    không nổi rồi báo file hỏng và đòi sửa. Ở đây toàn văn xuôi, không ô nào là
    công thức thật, nên ép hết về kiểu chuỗi trước khi lưu.
    """
    for ws in wb:
        for row in ws.iter_rows():
            for c in row:
                if c.data_type == "f":
                    c.data_type = "s"


def workbook_mot_bac(ver, tung_dong, bo_sot, tong_ket):
    """
    Gộp 3 file chấm tay của MỘT bậc thành một workbook đặt ngay trong out/<ver>/.
    3 file CSV vẫn giữ: script 2 ghi và script 3 đọc chúng, và git diff được CSV
    chứ không diff được xlsx — mất CSV là mất khả năng soi lịch sử sửa nhãn.
    """
    wb = Workbook()
    wb.remove(wb.active)
    tab_tung_dong(wb, [r for r in tung_dong if r[0] == ver])
    tab_bo_sot(wb, [r for r in bo_sot if r[0] == ver])
    tab_tong_ket(wb, [r for r in tong_ket if r[0] == ver])
    chan_cong_thuc(wb)

    dich = OUT / ver / f"NGUOI_CHAM_{ver}.xlsx"
    wb.save(dich)
    return dich


def main() -> int:
    may = doc(OUT / "may_cham_4_ban.csv")
    nguoi = {v: doc(OUT / v / "nguoi_cham_tong_ket.csv") for v in VERS}
    may_theo_tin = {v: doc(OUT / v / "may_cham.csv") for v in VERS}

    # Ba bảng thô của tầng người, giữ nguyên nội dung file CSV, chỉ thêm cột "bậc"
    # để 4 thư mục nằm chung một tab mà vẫn lọc riêng ra được.
    tung_dong = [(v, r) for v in VERS for r in doc(OUT / v / "nguoi_cham_tung_dong.csv")]
    bo_sot = [(v, r) for v in VERS for r in doc(OUT / v / "nguoi_cham_bo_sot.csv")]
    tong_ket = [(v, r) for v in VERS for r in nguoi[v]]

    wb = Workbook()
    wb.remove(wb.active)
    tab_huong_dan(wb)
    tab_tong_hop(wb, may, nguoi)
    tab_theo_tin(wb, nguoi, may_theo_tin)
    tab_tung_dong(wb, tung_dong)
    tab_bo_sot(wb, bo_sot)
    tab_tong_ket(wb, tong_ket)

    chan_cong_thuc(wb)
    dich = OUT / "KET_QUA_TONG_HOP.xlsx"
    wb.save(dich)
    print(f"Đã ghi: {dich}")
    print(f"  DocTruoc            — ai chấm cái gì, file nào ra số nào")
    print(f"  TongHop             — 4 bậc prompt, máy + người cạnh nhau")
    print(f"  TheoTin             — 10 tin × 4 bậc, dạng bảng ngang")
    print(f"  NguoiCham_TungDong  — {len(tung_dong)} dòng nhãn tay")
    print(f"  NguoiCham_BoSot     — {len(bo_sot)} dòng đếm bỏ sót")
    print(f"  NguoiCham_TongKet   — {len(tong_ket)} dòng P/R/F1")

    print("\nMỗi bậc thêm một workbook 3 tab ngay trong thư mục của nó:")
    for v in VERS:
        print(f"  {workbook_mot_bac(v, tung_dong, bo_sot, tong_ket)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
